"""Seed master data from the TASC sample xlsx into the DB.

Also seeds Contract / RateCard / SOW per-client so the BTP-style validation
rule engine has the per-jurisdiction config it needs (brief §4.1 & §4.5).

Run: `uv run python -m tia_ai.seed`
"""

from __future__ import annotations

import openpyxl

from .config import SEED_XLSX
from .db import SessionLocal, engine, init_db
from .models import Base, Client, Employee, Payroll


def _rows(ws):
    it = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(it)]
    for row in it:
        if all(c is None for c in row):
            continue
        yield dict(zip(header, row, strict=False))


def _num(v) -> float:
    try:
        return float(v) if v is not None else 0.0
    except (TypeError, ValueError):
        return 0.0


def get_session():
    # local accessor so the rest of this module reads naturally; mirrors db.get_session
    from .db import get_session as _gs

    return _gs()


def seed() -> dict[str, int]:
    # Drop+create so new columns/tables (Contract, RateCard, SOW, Query +
    # extended Invoice fields) are picked up on every reseed.
    Base.metadata.drop_all(engine)
    Base.metadata.create_all(engine)
    wb = openpyxl.load_workbook(SEED_XLSX, read_only=True, data_only=True)
    counts = {"clients": 0, "employees": 0, "payroll": 0}

    with get_session() as s:
        for r in _rows(wb["Customers"]):
            s.add(
                Client(
                    code=r["Client Code"],
                    name=r["Client Name"],
                    city=r.get("City"),
                    industry=r.get("Industry"),
                    contact_email=r.get("Contact Email"),
                    status=r.get("Status", "Active"),
                )
            )
            counts["clients"] += 1

        for r in _rows(wb["Employees"]):
            s.add(
                Employee(
                    emp_id=r["Emp ID"],
                    full_name=r["Full Name"],
                    first_name=r.get("First Name"),
                    last_name=r.get("Last Name"),
                    email=r.get("Email"),
                    client_code=r["Client Code"],
                    client_name=r.get("Client Name"),
                    job_title=r.get("Job Title"),
                    department=r.get("Department"),
                    nationality=r.get("Nationality"),
                    date_of_joining=str(r.get("Date of Joining")),
                    status=r.get("Status", "Active"),
                    iban=r.get("IBAN"),
                    basic=_num(r.get("Basic")),
                    housing=_num(r.get("Housing")),
                    transport=_num(r.get("Transport")),
                    food=_num(r.get("Food")),
                    phone=_num(r.get("Phone")),
                    total_ctc=_num(r.get("Total CTC")),
                )
            )
            counts["employees"] += 1

        payroll_sheet = next((n for n in wb.sheetnames if n.lower().startswith("payroll")), None)
        if payroll_sheet:
            for r in _rows(wb[payroll_sheet]):
                s.add(
                    Payroll(
                        emp_id=r["Emp ID"],
                        employee_name=r.get("Employee Name"),
                        client_code=r["Client Code"],
                        period=str(r.get("Pay Period")),
                        basic=_num(r.get("Basic")),
                        housing=_num(r.get("Housing")),
                        transport=_num(r.get("Transport")),
                        food=_num(r.get("Food")),
                        phone=_num(r.get("Phone")),
                        gross=_num(r.get("Gross")),
                        ot_hours=_num(r.get("OT Hours")),
                        ot_amount=_num(r.get("OT Amount")),
                        deductions=_num(r.get("Deductions")),
                        net_pay=_num(r.get("Net Pay")),
                        currency=r.get("Currency", "AED"),
                        working_days=int(_num(r.get("Working Days"))),
                    )
                )
                counts["payroll"] += 1

    # contracts depend on master data
    from .seed_contracts import seed_contracts

    counts.update(seed_contracts())
    return counts


if __name__ == "__main__":
    print("Seeding from", SEED_XLSX)
    print(seed())
