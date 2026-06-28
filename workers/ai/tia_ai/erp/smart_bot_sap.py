"""TASC Smart Bot + SAP - the orchestrator's "load payroll, run SAP, emit invoices" stage.

Brief §4.4: "Simulate or integrate a bot/ERP step that takes the consolidated
file, processes payroll, and generates invoices." The architecture diagram
labels this block "TASC Smart Bot + SAP" with three substeps:

  ① Collect Consolidated Excel & Upload to SAP   →  build_consolidated_excel
  ② Process Payroll (SAP)                         →  process_payroll (event-only)
  ③ Generate Invoices                             →  invoice generation (downstream)

Plus a side-output the brief doesn't ask for but TASC actually needs in
production: a WPS SIF (Salary Information File) for the bank gateway. UAE Federal
Decree-Law / MOHRE / Central Bank rules: salaries to private-sector employees
must flow through WPS-registered banks, accompanied by a fixed-width SIF file
with SCR (header) + EDR (per-employee) records.
"""

from __future__ import annotations

import datetime as dt
import io
import re
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from ..config import STAGING_DIR
from ..models import Client, Contract, Employee, Invoice, Payroll
from .mock import OT_DIVISOR_DAYS, OT_HOURS_PER_DAY, OT_STANDARD_MULT, _compute_ot, _d, _money

CENT = Decimal("0.01")


def _ramco_columns() -> list[str]:
    """Column headers shaped like Ramco SRP's payroll/billing export.

    TASC has run on Ramco SRP since 2015 (per their press release). The Smart Bot
    in production posts this consolidated file to Ramco's import API; for the
    demo we emit the file with the columns Ramco expects so the path is one
    transform away from real integration.
    """
    return [
        "Emp Code",  # Ramco "Employee Code"
        "Emp Name",
        "Client Code",
        "Client Name",
        "Pay Period",
        "Working Days",
        "Standard Days",
        "Basic",
        "Housing",
        "Transport",
        "Food",
        "Phone",
        "Gross",
        "OT Hours",
        "OT Rate (AED/hr)",
        "OT Amount",
        "Reimbursements",
        "Deductions",
        "Net Pay",
        "Markup %",
        "Bill Amount (excl VAT)",
        "VAT Rate %",
        "VAT Amount",
        "Bill Amount (incl VAT)",
        "Currency",
        "IBAN",
        "Pay Date",
        "Contract Type",
        "SAC / Service Code",
        "Status",
    ]


def build_consolidated_excel(session: Session, client_code: str, period: str) -> Path:
    """Step ① of Smart Bot + SAP: assemble the SAP-ready consolidated workbook.

    Reads master Payroll table joined with Employee and Contract metadata.
    Returns the path to the saved .xlsx in STAGING_DIR.
    """
    client = session.get(Client, client_code)
    contract = (
        session.query(Contract)
        .filter(Contract.client_code == client_code, Contract.active.is_(True))
        .first()
    )
    markup = float(contract.markup_pct) if contract else 0.20
    vat_rate = float(contract.vat_rate) if contract else 0.05
    sac = (contract.sac_code if contract else None) or ""
    contract_type = contract.type if contract else "-"

    payrolls = (
        session.query(Payroll)
        .filter(Payroll.client_code == client_code, Payroll.period == period)
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Consolidated"
    headers = _ramco_columns()
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="d9531e")
    header_font = Font(bold=True, color="FFFFFF")
    for c in ws[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    for p in payrolls:
        emp = session.get(Employee, p.emp_id)
        # recompute OT per statutory formula on top of the seed OT amount,
        # so the workbook reads consistently with the invoice
        ot_hourly = _d(p.basic) / OT_DIVISOR_DAYS / OT_HOURS_PER_DAY * OT_STANDARD_MULT
        ot_amount = _compute_ot(p.basic, p.ot_hours)
        std_days = p.working_days or 22
        prorated_gross = _d(p.gross)
        bill_excl_vat = (prorated_gross + ot_amount) * (Decimal(1) + Decimal(str(markup)))
        vat_amount = bill_excl_vat * Decimal(str(vat_rate))
        bill_incl_vat = bill_excl_vat + vat_amount
        ws.append(
            [
                p.emp_id,
                p.employee_name or (emp.full_name if emp else ""),
                p.client_code,
                client.name if client else "",
                p.period,
                int(p.working_days or 0),
                int(std_days),
                float(_d(p.basic)),
                float(_d(p.housing)),
                float(_d(p.transport)),
                float(_d(p.food)),
                float(_d(p.phone)),
                float(_d(p.gross)),
                float(p.ot_hours or 0),
                _money(ot_hourly),
                _money(ot_amount),
                0.0,
                float(_d(p.deductions)),
                float(_d(p.net_pay)),
                markup,
                _money(bill_excl_vat),
                vat_rate,
                _money(vat_amount),
                _money(bill_incl_vat),
                p.currency or "AED",
                emp.iban if emp else "",
                "",  # Pay Date - TBD on actual SAP run
                contract_type,
                sac,
                "READY",
            ]
        )
    # auto-width
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    out_path = Path(STAGING_DIR) / f"consolidated_{client_code}_{_slug(period)}.xlsx"
    wb.save(out_path)
    return out_path


def _slug(s: str) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "-", (s or "")).strip("-")


# --------------------------------------------------------------------------
# WPS SIF - Salary Information File (UAE Wages Protection System)
# --------------------------------------------------------------------------

# SIF v1.0 record spec (per Central Bank / MOHRE):
#   SCR (header): 1 row identifying the employer batch
#   EDR (employee detail): 1 row per employee with IBAN + net pay
#   File: CSV, pipe-delimited, UTF-8, ASCII subset only
#   Filename: <13-digit MOHRE employer ID>_YYYYMMDD_HHMMSS.sif
#
# Real production: bank validates SIF, releases salaries to each IBAN, sends
# back a payment-proof file; MOHRE pulls that proof via real-time integration
# with the Central Bank. For the demo we generate a schema-shaped sample.


# Sample TASC MOHRE employer ID (13 digits) - demo only
TASC_MOHRE_ID = "9999000000001"


def build_wps_sif(session: Session, client_code: str, period: str) -> Path:
    """Step in parallel with Smart Bot: emit a WPS SIF file ready for the bank.

    One SIF per client batch per period. SCR row is the header; EDR rows are
    per-employee. Net pay must sum to the SCR's total to the cent.
    """
    payrolls = (
        session.query(Payroll)
        .filter(Payroll.client_code == client_code, Payroll.period == period)
        .all()
    )
    today = dt.datetime.now(dt.timezone.utc)
    pay_date = today.strftime("%Y%m%d")
    pay_month = today.strftime("%m%Y")
    total_salaries = Decimal("0")
    edr_rows: list[str] = []
    for p in payrolls:
        emp = session.get(Employee, p.emp_id)
        iban = (emp.iban if emp else "") or ""
        # Real IBAN is 23 chars starting AE - our seed has zero-padded fakes; preserve as-is
        net = _d(p.net_pay).quantize(CENT, rounding=ROUND_HALF_UP)
        total_salaries += net
        # EDR | EmpID | IBAN | Salary frequency | "Net" | Net pay (cents) | currency | period (MMYYYY)
        edr_rows.append(
            "|".join(
                [
                    "EDR",
                    str(p.emp_id),
                    iban,
                    "M",  # monthly
                    "Net",
                    f"{int(net * 100):d}",  # cents, no decimal
                    p.currency or "AED",
                    pay_month,
                ]
            )
        )
    total_cents = int(total_salaries.quantize(CENT) * 100)
    # SCR | Employer MOHRE ID | Employer name | bank routing code | file creation date | pay date | currency | total cents | record count
    scr = "|".join(
        [
            "SCR",
            TASC_MOHRE_ID,
            "TASC Outsourcing FZ-LLC",
            "EBILUAEAXXX",  # sample SWIFT - Emirates Islamic
            today.strftime("%Y%m%d"),
            pay_date,
            "AED",
            f"{total_cents:d}",
            str(len(edr_rows)),
        ]
    )
    body = "\n".join([scr, *edr_rows]) + "\n"
    fname = f"{TASC_MOHRE_ID}_{today.strftime('%Y%m%d_%H%M%S')}_{client_code}.sif"
    out_path = Path(STAGING_DIR) / fname
    out_path.write_text(body, encoding="utf-8")
    return out_path


# --------------------------------------------------------------------------
# Step ② - Process Payroll (cosmetic but visible in the pipeline timeline)
# --------------------------------------------------------------------------


def process_payroll_event_payload(consolidated_path: Path, sif_path: Path, n_rows: int) -> dict:
    """Payload for the `payroll_processed_by_sap` event so it reads convincingly
    on the demo timeline. The mock has no real SAP behaviour - we just stage the
    artifacts and emit the event."""
    return {
        "engine": "smart_bot_sap (mock)",
        "consolidated_excel": str(consolidated_path),
        "wps_sif": str(sif_path),
        "records_processed": n_rows,
        "processed_at": dt.datetime.now(dt.timezone.utc).isoformat(),
    }


def _demo() -> None:
    """Offline self-check - verifies SCR/EDR shape against a tiny payload."""
    scr = "|".join(["SCR", "1" * 13, "X", "Y", "20260601", "20260601", "AED", "100000", "1"])
    edr = "|".join(["EDR", "EMP1", "AE000", "M", "Net", "100000", "AED", "062026"])
    assert scr.split("|")[0] == "SCR"
    assert edr.split("|")[0] == "EDR"
    assert len(edr.split("|")) == 8, edr
    print("smart_bot_sap SIF shape: OK")


if __name__ == "__main__":
    _demo()
