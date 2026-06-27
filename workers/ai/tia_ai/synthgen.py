"""Generate the 7 sample-input cases + eval gold from the brief.

Writes inputs to data/synthetic/ and ground truth to data/gold/.
Run: `uv run python -m tia_ai.synthgen`
"""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl
from openpyxl import Workbook

from .config import DATA_DIR

SYN = DATA_DIR / "synthetic"
GOLD = DATA_DIR / "gold"
SYN.mkdir(parents=True, exist_ok=True)
GOLD.mkdir(parents=True, exist_ok=True)

PERIOD = "June 2026"


def _write(path: Path, text: str) -> None:
    path.write_text(text, encoding="utf-8")


def _gold(case: str, expect: dict) -> None:
    (GOLD / f"case_{case}.json").write_text(json.dumps(expect, indent=2), encoding="utf-8")


# ---------------------------------------------------------------- case 7
def case07_clean_excel() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Timesheet"
    ws.append(["Emp ID", "Full Name", "Client Code", "Period", "Working Days", "OT Hours", "Leave"])
    rows = [
        ("EMP10001", "Carlos Smith", "CL001", PERIOD, 22, 2, ""),
        ("EMP10002", "Ahmed Khan", "CL001", PERIOD, 20, 4, "AL"),
        ("EMP10003", "Meera Al Rashid", "CL001", PERIOD, 21, 0, ""),
    ]
    for r in rows:
        ws.append(r)
    wb.save(SYN / "case_07_clean.xlsx")
    _gold(
        "07",
        {
            "case": "07",
            "channel": "upload",
            "input": "case_07_clean.xlsx",
            "expect": {
                "client_code": "CL001",
                "period": PERIOD,
                "rows": [
                    {
                        "emp_id": "EMP10001",
                        "employee_name": "Carlos Smith",
                        "days_worked": 22,
                        "ot_hours": 2,
                        "resolved": True,
                        "ambiguous": False,
                    },
                    {
                        "emp_id": "EMP10002",
                        "employee_name": "Ahmed Khan",
                        "days_worked": 20,
                        "ot_hours": 4,
                        "resolved": True,
                        "ambiguous": False,
                    },
                    {
                        "emp_id": "EMP10003",
                        "employee_name": "Meera Al Rashid",
                        "days_worked": 21,
                        "ot_hours": 0,
                        "resolved": True,
                        "ambiguous": False,
                    },
                ],
            },
        },
    )


# ---------------------------------------------------------------- case 5
def case05_punch_excel() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Punch"
    # 5 working days shown as In/Out pairs + a leave column
    days = 5
    header = ["Emp ID", "Full Name", "Client Code", "Period"]
    for d in range(1, days + 1):
        header += [f"D{d} In", f"D{d} Out"]
    header += ["Leave"]
    ws.append(header)

    def punch_row(emp, name, present_days, leave):
        row = [emp, name, "CL001", PERIOD]
        for d in range(1, days + 1):
            if d <= present_days:
                row += ["09:00", "17:00"]
            else:
                row += ["", ""]
        row += [leave]
        return row

    ws.append(punch_row("EMP10001", "Carlos Smith", 5, ""))
    ws.append(punch_row("EMP10002", "Ahmed Khan", 4, "A/L"))  # mixed leave spelling
    ws.append(punch_row("EMP10003", "Meera Al Rashid", 3, "sick"))
    wb.save(SYN / "case_05_punch.xlsx")
    _gold(
        "05",
        {
            "case": "05",
            "channel": "upload",
            "input": "case_05_punch.xlsx",
            "expect": {
                "client_code": "CL001",
                "period": PERIOD,
                "rows": [
                    {
                        "emp_id": "EMP10001",
                        "employee_name": "Carlos Smith",
                        "days_worked": 5,
                        "hours": 40.0,
                        "resolved": True,
                        "ambiguous": False,
                    },
                    {
                        "emp_id": "EMP10002",
                        "employee_name": "Ahmed Khan",
                        "days_worked": 4,
                        "hours": 32.0,
                        "leave_codes": ["AL"],
                        "resolved": True,
                        "ambiguous": False,
                    },
                    {
                        "emp_id": "EMP10003",
                        "employee_name": "Meera Al Rashid",
                        "days_worked": 3,
                        "hours": 24.0,
                        "leave_codes": ["SICK"],
                        "resolved": True,
                        "ambiguous": False,
                    },
                ],
            },
        },
    )


# ---------------------------------------------------------------- case 2
def case02_email_employee() -> None:
    body = f"""Subject: My timesheet for {PERIOD}

Hi Payroll team,

My employee id is EMP10001 and I worked 22 days this month with 2 OT hours.

Regards,
Carlos
"""
    _write(SYN / "case_02_email_employee.txt", body)
    _gold(
        "02",
        {
            "case": "02",
            "channel": "email",
            "input": "case_02_email_employee.txt",
            "expect": {
                "period": PERIOD,
                "rows": [
                    {
                        "emp_id": "EMP10001",
                        "employee_name": "Carlos Smith",
                        "days_worked": 22,
                        "ot_hours": 2,
                        "resolved": True,
                        "ambiguous": False,
                    },
                ],
            },
        },
    )


# ---------------------------------------------------------------- case 3
def case03_email_client_full() -> None:
    body = f"""Subject: Monthly timesheet submission

Client: Emirates Steel Industries LLC
Period: {PERIOD}

Carlos Smith - 22 days
Ahmed Khan - 20 days, 4 OT hours
Meera Al Rashid - 21 days

Approved by: Site Manager
"""
    _write(SYN / "case_03_email_client_full.eml", body)
    _gold(
        "03",
        {
            "case": "03",
            "channel": "email",
            "input": "case_03_email_client_full.eml",
            "expect": {
                "client_code": "CL001",
                "period": PERIOD,
                "rows": [
                    {
                        "employee_name": "Carlos Smith",
                        "days_worked": 22,
                        "resolved": True,
                        "ambiguous": False,
                    },
                    {
                        "employee_name": "Ahmed Khan",
                        "days_worked": 20,
                        "ot_hours": 4,
                        "resolved": True,
                        "ambiguous": False,
                    },
                    {
                        "employee_name": "Meera Al Rashid",
                        "days_worked": 21,
                        "resolved": True,
                        "ambiguous": False,
                    },
                ],
            },
        },
    )


# ---------------------------------------------------------------- case 6
def case06_email_structured() -> None:
    body = f"""Subject: Leave and reimbursements {PERIOD}

Client: Emirates Steel Industries LLC
Period: {PERIOD}

EMP10001 Carlos Smith - 20 days, leave: AL, reimbursement AED 250 for taxi
EMP10002 Ahmed Khan - 22 days, claim AED 120 for parking

Regards,
Finance
"""
    _write(SYN / "case_06_email_structured.txt", body)
    _gold(
        "06",
        {
            "case": "06",
            "channel": "email",
            "input": "case_06_email_structured.txt",
            "expect": {
                "client_code": "CL001",
                "period": PERIOD,
                "rows": [
                    {
                        "emp_id": "EMP10001",
                        "employee_name": "Carlos Smith",
                        "days_worked": 20,
                        "leave_codes": ["AL"],
                        "reimbursements": 1,
                        "resolved": True,
                        "ambiguous": False,
                    },
                    {
                        "emp_id": "EMP10002",
                        "employee_name": "Ahmed Khan",
                        "days_worked": 22,
                        "reimbursements": 1,
                        "resolved": True,
                        "ambiguous": False,
                    },
                ],
            },
        },
    )


# ---------------------------------------------------------------- case 1 (ambiguous)
def case01_email_no_empid() -> None:
    body = f"""Subject: Payout request

Client: Majid Al Futtaim Retail LLC
Period: {PERIOD}

Fatima Khan - 23 days, total AED 12000

Regards,
Operations
"""
    _write(SYN / "case_01_email_no_empid.eml", body)
    _gold(
        "01",
        {
            "case": "01",
            "channel": "email",
            "input": "case_01_email_no_empid.eml",
            "expect": {
                "client_code": "CL005",
                "period": PERIOD,
                "rows": [
                    {
                        "employee_name": "Fatima Khan",
                        "days_worked": 23,
                        "resolved": False,
                        "ambiguous": True,
                        "candidate_emp_ids": ["EMP10083", "EMP10093"],
                    },
                ],
            },
        },
    )


# ---------------------------------------------------------------- case 4 (handwritten)
def case04_handwritten() -> None:
    from PIL import Image, ImageDraw, ImageFont

    img = Image.new("RGB", (900, 600), "white")
    d = ImageDraw.Draw(img)
    try:
        font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 26)
        small = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 20)
    except OSError:
        font = ImageFont.load_default()
        small = font
    d.text((40, 30), "TIMESHEET - Emirates Steel Industries LLC", fill="black", font=font)
    d.text((40, 70), f"Period: {PERIOD}", fill="black", font=small)
    lines = [
        "Carlos Smith        22 days   2 OT",
        "Ahmed Khan          20 days   AL",
        "Meera Al Rashid     21 days",
    ]
    y = 140
    for ln in lines:
        d.text((60, y), ln, fill="navy", font=small)
        y += 50
    d.text((60, 360), "Signed: Site Manager", fill="black", font=small)
    d.rectangle([560, 380, 840, 500], outline="red", width=3)
    d.text((580, 420), "CLIENT STAMP", fill="red", font=small)
    d.text((580, 450), "Emirates Steel", fill="red", font=small)
    img.save(SYN / "case_04_handwritten.png")
    _gold(
        "04",
        {
            "case": "04",
            "channel": "upload",
            "input": "case_04_handwritten.png",
            "expect": {
                "client_code": "CL001",
                "period": PERIOD,
                "rows": [
                    {
                        "employee_name": "Carlos Smith",
                        "days_worked": 22,
                        "ot_hours": 2,
                        "resolved": True,
                        "ambiguous": False,
                    },
                    {
                        "employee_name": "Ahmed Khan",
                        "days_worked": 20,
                        "leave_codes": ["AL"],
                        "resolved": True,
                        "ambiguous": False,
                    },
                    {
                        "employee_name": "Meera Al Rashid",
                        "days_worked": 21,
                        "resolved": True,
                        "ambiguous": False,
                    },
                ],
            },
        },
    )


def generate_all() -> list[str]:
    case01_email_no_empid()
    case02_email_employee()
    case03_email_client_full()
    case04_handwritten()
    case05_punch_excel()
    case06_email_structured()
    case07_clean_excel()
    case08_aisha_3way()
    case09_messy_excel()
    case10_email_quoted_reply()
    case11_clean_pdf()
    case12_rate_mismatch()
    case13_out_of_scope_sow()
    case14_ot_over_cap()
    return sorted(p.name for p in SYN.iterdir() if p.is_file() and p.name != ".gitkeep")


# ---------------------------------------------------------------- case 8 (3-way ambiguity, cross-client)
def case08_aisha_3way() -> None:
    body = f"""Subject: Timesheet for Aisha

Period: {PERIOD}

Aisha Al Zaabi - 22 days

Thanks
"""
    _write(SYN / "case_08_aisha_3way.eml", body)
    _gold(
        "08",
        {
            "case": "08",
            "channel": "email",
            "input": "case_08_aisha_3way.eml",
            "expect": {
                "period": PERIOD,
                "rows": [
                    {
                        "employee_name": "Aisha Al Zaabi",
                        "days_worked": 22,
                        "resolved": False,
                        "ambiguous": True,
                        "candidate_emp_ids": ["EMP10058", "EMP10072", "EMP10077"],
                    },
                ],
            },
        },
    )


# ---------------------------------------------------------------- case 9 (messy Excel)
def case09_messy_excel() -> None:
    """Real-world Excel pain: extra blank rows, an extra unrecognized column."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Timesheet"
    ws.append(
        [
            "Emp ID",
            "Full Name",
            "Client Code",
            "Period",
            "Working Days",
            "OT Hours",
            "Leave",
            "Notes",
        ]
    )
    ws.append(("EMP10001", "Carlos Smith", "CL001", PERIOD, 22, 2, "", "primary"))
    ws.append((None,) * 8)  # blank row
    ws.append(("EMP10002", "Ahmed Khan", "CL001", PERIOD, 20, 4, "AL", ""))
    ws.append(("EMP10003", "Meera Al Rashid", "CL001", PERIOD, 21, 0, "", "deputy"))
    ws.append((None,) * 8)  # trailing blank
    wb.save(SYN / "case_09_messy.xlsx")
    _gold(
        "09",
        {
            "case": "09",
            "channel": "upload",
            "input": "case_09_messy.xlsx",
            "expect": {
                "client_code": "CL001",
                "period": PERIOD,
                "rows": [
                    {
                        "emp_id": "EMP10001",
                        "employee_name": "Carlos Smith",
                        "days_worked": 22,
                        "ot_hours": 2,
                        "resolved": True,
                        "ambiguous": False,
                    },
                    {
                        "emp_id": "EMP10002",
                        "employee_name": "Ahmed Khan",
                        "days_worked": 20,
                        "ot_hours": 4,
                        "leave_codes": ["AL"],
                        "resolved": True,
                        "ambiguous": False,
                    },
                    {
                        "emp_id": "EMP10003",
                        "employee_name": "Meera Al Rashid",
                        "days_worked": 21,
                        "ot_hours": 0,
                        "resolved": True,
                        "ambiguous": False,
                    },
                ],
            },
        },
    )


# ---------------------------------------------------------------- case 10 (quoted-reply email)
def case10_email_quoted_reply() -> None:
    """Email with a quoted-reply thread underneath. The quoted history mentions
    different day counts (25 days). The parser must ignore quoted lines and only
    extract the rows from the latest message at the top."""
    body = f"""Subject: Re: June timesheet submission

Client: Emirates Steel Industries LLC
Period: {PERIOD}

Carlos Smith - 22 days
Ahmed Khan - 20 days, 4 OT hours

Thanks,
Finance team

> On Mon, 1 Jul 2026, Site Manager <manager@steel.test> wrote:
> Please submit your June 2026 timesheet by Friday.
> For reference, last month Carlos Smith logged 25 days.
> Ahmed Khan was on AL for 10 days.
> Best,
> Site Manager
"""
    _write(SYN / "case_10_email_quoted_reply.eml", body)
    _gold(
        "10",
        {
            "case": "10",
            "channel": "email",
            "input": "case_10_email_quoted_reply.eml",
            "expect": {
                "client_code": "CL001",
                "period": PERIOD,
                "rows": [
                    {
                        "employee_name": "Carlos Smith",
                        "days_worked": 22,
                        "resolved": True,
                        "ambiguous": False,
                    },
                    {
                        "employee_name": "Ahmed Khan",
                        "days_worked": 20,
                        "ot_hours": 4,
                        "resolved": True,
                        "ambiguous": False,
                    },
                ],
            },
        },
    )


# ---------------------------------------------------------------- case 11 (clean PDF — required §7.5 deliverable)
def case11_clean_pdf() -> None:
    """A typed/printed PDF with a real text layer — satisfies brief §7.5
    'sample inputs: Excel, PDF, handwritten'. Uses Typst so pdfplumber can
    extract text cleanly without OCR."""
    import typst

    src = f"""
#set page(paper: "a4", margin: 1.5cm)
#set text(size: 11pt)

#text(size: 18pt, weight: "bold")[Monthly Timesheet]

#v(6pt)
Client: Emirates Steel Industries LLC (CL001)

Period: {PERIOD}

#v(10pt)

EMP10001 Carlos Smith - 22 days, 2 OT hours

EMP10002 Ahmed Khan - 20 days, leave: AL

EMP10003 Meera Al Rashid - 21 days

#v(20pt)
Approved by: Site Manager
"""
    typ_path = SYN / "_case_11_clean.typ"
    pdf_path = SYN / "case_11_clean_pdf.pdf"
    typ_path.write_text(src, encoding="utf-8")
    typst.compile(str(typ_path), output=str(pdf_path))
    typ_path.unlink(missing_ok=True)
    _gold(
        "11",
        {
            "case": "11",
            "channel": "upload",
            "input": "case_11_clean_pdf.pdf",
            "expect": {
                "client_code": "CL001",
                "period": PERIOD,
                "rows": [
                    {
                        "emp_id": "EMP10001",
                        "employee_name": "Carlos Smith",
                        "days_worked": 22,
                        "ot_hours": 2,
                        "resolved": True,
                        "ambiguous": False,
                    },
                    {
                        "emp_id": "EMP10002",
                        "employee_name": "Ahmed Khan",
                        "days_worked": 20,
                        "leave_codes": ["AL"],
                        "resolved": True,
                        "ambiguous": False,
                    },
                    {
                        "emp_id": "EMP10003",
                        "employee_name": "Meera Al Rashid",
                        "days_worked": 21,
                        "resolved": True,
                        "ambiguous": False,
                    },
                ],
            },
        },
    )


# ---------------------------------------------------------------- case 12 (rate mismatch — rule R2 fires)
def case12_rate_mismatch() -> None:
    """Client bills Carlos Smith at AED 300/hr but contract rate card says 225/hr."""
    body = f"""Subject: June timesheet — special rate request

Client: Emirates Steel Industries LLC (CL001)
Period: {PERIOD}

EMP10001 Carlos Smith - 22 days, 2 OT hours
Billing rate: AED 300/hr (please honor this rate per Q1 agreement)

Best,
Site Manager
"""
    _write(SYN / "case_12_rate_mismatch.eml", body)
    _gold(
        "12",
        {
            "case": "12",
            "channel": "email",
            "input": "case_12_rate_mismatch.eml",
            "expect": {
                "client_code": "CL001",
                "period": PERIOD,
                "rows": [
                    {
                        "emp_id": "EMP10001",
                        "employee_name": "Carlos Smith",
                        "days_worked": 22,
                        "ot_hours": 2,
                        "resolved": True,
                        "ambiguous": False,
                    },
                ],
                # contract-bound rule R2 should fire (billed rate exceeds rate-card regular_rate)
                "expected_routing": "hitl",
                "expected_rule_violation": "R2_rate_compliance_per_category",
            },
        },
    )


# ---------------------------------------------------------------- case 13 (out-of-scope hours — rule R5 fires)
def case13_out_of_scope_sow() -> None:
    """CL002 Emaar has FIXED_SCOPE contract; 'Design phase' SOW is COMPLETED. Bill hours
    against that closed deliverable to trigger R5."""
    body = f"""Subject: June work — Design phase continuation

Client: Emaar Properties PJSC (CL002)
Period: {PERIOD}
SOW: Design phase

EMP10021 - 22 days, 16 OT hours (design rework as agreed)

The team continued refining the design through June.

Regards,
Operations
"""
    _write(SYN / "case_13_out_of_scope_sow.eml", body)
    _gold(
        "13",
        {
            "case": "13",
            "channel": "email",
            "input": "case_13_out_of_scope_sow.eml",
            "expect": {
                "client_code": "CL002",
                "period": PERIOD,
                "rows": [
                    {
                        "emp_id": "EMP10021",
                        "days_worked": 22,
                        "ot_hours": 16,
                        "resolved": True,
                        "ambiguous": False,
                    },
                ],
                "expected_routing": "hitl",
                "expected_rule_violation": "R5_sow_hours_not_exceeded",
            },
        },
    )


# ---------------------------------------------------------------- case 14 (OT over contract cap — rule R4 fires)
def case14_ot_over_cap() -> None:
    """OT 50 hours on top of 22×8=176 regular hours = 28% — exceeds contract max_ot_pct=20%."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Timesheet"
    ws.append(["Emp ID", "Full Name", "Client Code", "Period", "Working Days", "OT Hours", "Leave"])
    ws.append(("EMP10001", "Carlos Smith", "CL001", PERIOD, 22, 50, ""))
    wb.save(SYN / "case_14_ot_over_cap.xlsx")
    _gold(
        "14",
        {
            "case": "14",
            "channel": "upload",
            "input": "case_14_ot_over_cap.xlsx",
            "expect": {
                "client_code": "CL001",
                "period": PERIOD,
                "rows": [
                    {
                        "emp_id": "EMP10001",
                        "employee_name": "Carlos Smith",
                        "days_worked": 22,
                        "ot_hours": 50,
                        "resolved": True,
                        "ambiguous": False,
                    },
                ],
                "expected_routing": "hitl",
                "expected_rule_violation": "R4_ot_within_contract_cap",
            },
        },
    )


if __name__ == "__main__":
    files = generate_all()
    print("generated:", files)
