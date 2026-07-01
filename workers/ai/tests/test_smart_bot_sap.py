"""TASC Smart Bot + SAP stage (erp/smart_bot_sap.py).

Covers the consolidated Ramco-shaped Excel export, the WPS SIF file (SCR/EDR
records + cent-exact totals), and the payroll-processed event payload. Uses a
real seeded (client, period) that has payroll rows.
"""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from tia_ai.db import SessionLocal
from tia_ai.erp.smart_bot_sap import (
    TASC_MOHRE_ID,
    _ramco_columns,
    _slug,
    build_consolidated_excel,
    build_wps_sif,
    process_payroll_event_payload,
)
from tia_ai.models import Payroll


@pytest.fixture(scope="module")
def client_period():
    s = SessionLocal()
    try:
        pr = s.query(Payroll).filter(Payroll.gross > 0).first()
        assert pr is not None, "seed payroll missing"
        cc, period = pr.client_code, pr.period
        n = s.query(Payroll).filter(Payroll.client_code == cc, Payroll.period == period).count()
        return cc, period, n
    finally:
        s.close()


def test_consolidated_excel_headers_and_rows(client_period):
    cc, period, n = client_period
    s = SessionLocal()
    try:
        path = build_consolidated_excel(s, cc, period)
    finally:
        s.close()
    assert path.exists()
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert headers == _ramco_columns()
    # one data row per payroll record + the header row
    assert ws.max_row == n + 1
    assert "Bill Amount (incl VAT)" in headers


def test_consolidated_excel_vat_math_reconciles(client_period):
    cc, period, _ = client_period
    s = SessionLocal()
    try:
        path = build_consolidated_excel(s, cc, period)
    finally:
        s.close()
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    cols = {c.value: i for i, c in enumerate(ws[1])}
    bi = cols["Bill Amount (excl VAT)"]
    vi = cols["VAT Amount"]
    ti = cols["Bill Amount (incl VAT)"]
    ri = cols["VAT Rate %"]
    row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    excl, vat, incl, rate = row[bi], row[vi], row[ti], row[ri]
    # incl == excl + vat, and vat == excl * rate (to the cent)
    assert abs((excl + vat) - incl) < 0.01
    assert abs(vat - round(excl * rate, 2)) < 0.01


def test_wps_sif_structure_and_totals(client_period):
    cc, period, n = client_period
    s = SessionLocal()
    try:
        path = build_wps_sif(s, cc, period)
    finally:
        s.close()
    assert path.exists()
    assert path.name.startswith(TASC_MOHRE_ID)
    lines = path.read_text(encoding="utf-8").strip().splitlines()
    scr = lines[0].split("|")
    edrs = [ln.split("|") for ln in lines[1:]]
    assert scr[0] == "SCR"
    assert scr[1] == TASC_MOHRE_ID
    assert all(e[0] == "EDR" for e in edrs)
    assert len(edrs) == n
    # SCR record count field matches the number of EDR rows
    assert int(scr[-1]) == n
    # SCR total cents == sum of EDR net-pay cents (cent-exact WPS requirement)
    scr_total_cents = int(scr[7])
    edr_total_cents = sum(int(e[5]) for e in edrs)
    assert scr_total_cents == edr_total_cents


def test_wps_sif_amounts_match_payroll_net(client_period):
    cc, period, _ = client_period
    s = SessionLocal()
    try:
        path = build_wps_sif(s, cc, period)
        rows = s.query(Payroll).filter(Payroll.client_code == cc, Payroll.period == period).all()
        expected_cents = sum(
            int((Decimal(str(p.net_pay or 0)).quantize(Decimal("0.01"))) * 100) for p in rows
        )
    finally:
        s.close()
    scr_total = int(path.read_text().splitlines()[0].split("|")[7])
    assert scr_total == expected_cents


def test_process_payroll_event_payload_shape():
    payload = process_payroll_event_payload(Path("/tmp/cons.xlsx"), Path("/tmp/x.sif"), 42)
    assert payload["records_processed"] == 42
    assert payload["consolidated_excel"].endswith("cons.xlsx")
    assert payload["wps_sif"].endswith("x.sif")
    assert "processed_at" in payload


def test_slug_normalises_period():
    assert _slug("June 2026") == "June-2026"
    assert _slug("") == ""
